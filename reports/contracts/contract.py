from pandas import DataFrame as pd_DataFrame
from pandas import concat as pd_concat
from pandas import merge as pd_merge
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from io import BytesIO
from copy import copy
from common import full_path
import re
from json import load as js_load

def get_group_fields():
    revenue = {
        "filial": "filial",
        "dogovorssilka": "dogovorssilka",
        "doh_dogovor": "doh_dogovor",
        "doh_kontragent": "doh_kontragent",
        "inn": "inn",
        "doh_nomerdogovora": "doh_nomerdogovora",
        "doh_data": "doh_data",
        "doh_dataokonchaniya": "doh_dataokonchaniya",
        "doh_obektdogovora": "doh_obektdogovora",
        "tipdogovora": "tipdogovora",
        "summadogovora": "summadogovora",
        "summanachalo": "summanachalo",
        "ostatoknachalo": "ostatoknachalo",
        "oplatananachalo": "oplatananachalo",
        "summakonec": "summakonec",
        "ostatokkonec": "ostatokkonec"}

    expense = {
        "filial": "filial",
        "dogovorssilka": "dogovorssilka",
        "doh_dogovor": "doh_dogovor",
        "rash_dogovor": "rash_dogovor",
        "tipdogovora": "tipdogovora",
        "rash_kontragent": "rash_kontragent",
        "inn": "inn",
        "rash_nomerdogovora": "rash_nomerdogovora",
        "rash_data": "rash_data",
        "summadogovora": "summadogovora",
        "summanachalo": "summanachalo",
        "ostatoknachalo": "ostatoknachalo",
        "summakonec": "summakonec",
        "ostatokkonec": "ostatokkonec"}

    list_rev = list(revenue.values())
    for exp_var in list(expense.values()):
        if exp_var not in list_rev:
            list_rev.append(exp_var)

    return list_rev

def mapping_df_xls():
    dct = {
        'A': ('filial', ''),
        'B': ('doh_kontragent', ''),
        'C': (),
        'J': ('inn', ''),
        'D': ('doh_nomerdogovora', ''),
        'E': ('doh_data', ''),
        'F': ('doh_dataokonchaniya', ''),
        'G': ('doh_obektdogovora', ''),
        'H': ('tipdogovora', ''),
        'I': ('rash_kontragent', ''),
        'K': ('rash_nomerdogovora', ''),
        'L': ('rash_data', ''),
        'M': ('summadogovora', ''),
        'N': ('summanachalo', ''),
        'O': ('ostatoknachalo', ''),
        'P': ('oplatananachalo', '')
    }
    return dct

def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def get_quarter(month):
    return (int(month) - 1) // 3 + 1

def user_format(frm_value):
    pattern = re.compile("[0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]")
    ismatch = re.match(pattern, frm_value)
    if ismatch != None:
        date_str = frm_value[0:10].split("-")
        result_frm = f"{get_quarter(date_str[1])} кв. {date_str[0]} г."
    elif frm_value == "All":
        result_frm = "Всего"
    else:
        result_frm = frm_value
    return result_frm

def df_to_excel(table, template="", columns=None):
    lw = load_workbook(template)
    lw_sheet = lw.active
    """dynamic column starts"""
    dyn_col = 17
    """header style"""
    header_style = copy(lw_sheet["M2"]._style)
    header_num_style = copy(lw_sheet["M3"]._style)
    fill_pattern_blue = PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid')
    header_need = True
    """resources """
    resources = {
        "summazaperiodplanitogo": "Выручка\nплан.",
        "viruchkazaperiodfakt": "Выручка\nфакт",
        "oplatazaperiodplan": "Оплата\nплан",
        "oplatazaperiodfakt": "Оплата\nфакт"
    }
    """end"""
    end_colm = {
        ('summakonec', ''): 'Выполнено на конец',
        ('ostatokkonec', ''): 'Остаток не сданных работ на конец'
    }

    resource_column = dyn_col
    grouping_clm = dyn_col
    grouping_list = []
    column_for_rows_end = 0
    row_start = 4
    last_row = len(table.index) + row_start - 1
    grouping_row = row_start + 1
    for df_row in table.iterrows():
        """define revenue contract"""
        is_rev = df_row[0][1] == ""
        """static fields"""
        for key, value in columns.items():
            new_row = lw_sheet[f"{key}{row_start}"]
            df_curr_row = df_row[1]
            if is_rev:
                new_row.fill = fill_pattern_blue
            new_row.number_format = 'General'
            if value != ():
                new_row.value = df_curr_row[value]

        lw_sheet[f"A{row_start}"].comment = Comment(f"{df_row[0][0]}\n{df_row[0][1]}", "user", height=50, width=350)

        """pivot and the end"""
        """header"""
        if header_need:
            for key, value in resources.items():
                for idx in df_row[1][key].items():
                    """header"""
                    new_header_cell = lw_sheet.cell(2, resource_column)
                    new_header_cell.value = f"{value}\n{user_format(idx[0])}"
                    new_header_cell._style = header_style
                    """header number"""
                    new_num_header_cell = lw_sheet.cell(3, resource_column)
                    new_num_header_cell.value = resource_column
                    new_num_header_cell._style = header_num_style
                    resource_column += 1
                """grouping totals"""
                num_added_clm = (resource_column - grouping_clm)
                if num_added_clm > 1:
                    grouping_list.append((colnum_string(grouping_clm), colnum_string(grouping_clm + num_added_clm - 2)))
                    grouping_clm = grouping_clm + num_added_clm
        """rows"""
        column_for_rows = dyn_col
        for key in resources:
            for idx in df_row[1][key].items():
                new_row_cell = lw_sheet.cell(row_start, column_for_rows)
                new_row_cell.value = idx[1]
                new_row_cell.number_format = 'General'
                if is_rev:
                    new_row_cell.fill = fill_pattern_blue
                column_for_rows += 1
        """end"""

        """header"""
        if header_need:
            column_for_rows_end = resource_column
            for key, value in end_colm.items():
                new_header_cell = lw_sheet.cell(2, resource_column)
                new_header_cell.value = value
                new_header_cell._style = header_style
                """header number"""
                new_num_header_cell = lw_sheet.cell(3, resource_column)
                new_num_header_cell.value = resource_column
                new_num_header_cell._style = header_num_style
                resource_column += 1
        """rows"""
        column_for_rows = column_for_rows_end
        for key in list(end_colm.keys()):
            new_row_cell = lw_sheet.cell(row_start, column_for_rows)
            df_curr_row = df_row[1]
            new_row_cell.value = df_curr_row[key]
            new_row_cell.number_format = 'General'
            if is_rev:
                new_row_cell.fill = fill_pattern_blue
            column_for_rows += 1

        header_need = False
        """grouping row"""
        if is_rev:
            lw_sheet.row_dimensions.group(grouping_row, row_start - 1)
            grouping_row = row_start + 1
        elif row_start == last_row:
            lw_sheet.row_dimensions.group(grouping_row, row_start)
        row_start += 1
    """other styles"""
    """column size"""
    lw_sheet.auto_filter.ref = lw_sheet.dimensions
    for col_num in range(dyn_col, resource_column):
        lw_sheet.column_dimensions[colnum_string(col_num)].width = 14
    """fix A2"""
    lw_sheet.freeze_panes = lw_sheet.cell(3, 1)
    """grouping"""
    for grpcol in grouping_list:
        lw_sheet.column_dimensions.group(grpcol[0], grpcol[1])

    bin_report = BytesIO()
    # lw.save(full_path("reports/contracts/files/result.xlsx"))
    lw.save(bin_report)

    return BytesIO(bin_report.getvalue())

def get_contract_report(data_load):
    # with open(full_path("reports/contracts/files/full.json"), "r", encoding='utf-8') as js_file:
    #     data_load = js_load(js_file)

    df_period = pd_DataFrame(data_load["dannieperiod"])
    df_rev = pd_DataFrame(data_load["dohdogovori"])
    df_exp = pd_DataFrame(data_load["rashdogovori"])

    """union and merge table"""
    df_result = pd_concat([df_rev, df_exp], ignore_index=True, sort=False)
    # df_result = df_result.loc[df_result['doh_dogovor'] == '"Договор 00000006249 от 09.08.2019 12:00:01"']
    df_result = pd_merge(df_result, df_period, on=["dogovorssilka", "filial"], how="left")
    df_result = df_result.fillna(
        {"period": "2019-01-01T00:00:00", "oplatazaperiodplan": 0, "oplatazaperiodfakt": 0, "summazaperiodplanitogo": 0,
         "viruchkazaperiodfakt": 0}).fillna('')
    """building pivot table"""
    df_result = df_result.pivot_table(index=get_group_fields(), columns="period",
                                      values=["oplatazaperiodplan", "oplatazaperiodfakt", "summazaperiodplanitogo",
                                              "viruchkazaperiodfakt"], fill_value='',
                                      margins=True,
                                      aggfunc=sum)
    """removing the last row with totals"""
    df_result = df_result[:-1]
    """to swipe columns"""
    # result = result.swaplevel(0, 1, axis=1).sort_index(axis=1)
    """reseting and building a new hierarchy indexes"""
    df_result = df_result.reset_index()
    df_result = df_result.set_index(['doh_dogovor', 'rash_dogovor']).sort_index()

    """format values(need to rebuild later)"""
    df_result['doh_data'] = df_result['doh_data'].apply(lambda x: "{:.10}".format(x))
    df_result['doh_dataokonchaniya'] = df_result['doh_dataokonchaniya'].apply(lambda x: "{:.10}".format(x))
    df_result['rash_data'] = df_result['rash_data'].apply(lambda x: "{:.10}".format(x))
    """replacing 0 with empty string for excel"""
    df_result = df_result.replace(0, '')
    """creating excel file"""
    return df_to_excel(df_result, full_path("reports/contracts/files/contract_sketch.xlsx"), mapping_df_xls())

# rslt = get_contract_report()
